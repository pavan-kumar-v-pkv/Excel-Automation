"""
Summary Builder Module
Generates summary sheets from data sheets
"""

from typing import Dict, List, Optional, Tuple
import openpyxl.styles

class SummaryBuilder:
    """
    Build summary sheet from data sheets
    """

    def __init__(self, workbook, config=None, data_workbook=None):
        """
        Initialize summary builder

        Args:
            workbook: openpyxl workbook object
            config: Configuration object
            data_workbook: openpyxl workbook opened with data_only=True for formula results
        """
        from .config import Config

        self.workbook = workbook
        self.data_workbook = data_workbook
        self.config = config or Config()
        self.summary_data = []

    def build_summary(self) -> bool:
        """
        Build or update summary sheet

        Returns:
            True if successful
        """
        print(f"\n📊 Building Summary Sheet...")

        # Collect data from all sheets
        self._collect_sheet_totals()

        if not self.summary_data:
            print("    No data found to summarise")
            return False
        
        # get or create summary sheet
        summary_sheet = self._get_summary_sheet()

        # Write summary data
        self._write_summary(summary_sheet)

        print(f"Summary sheet updated with {len(self.summary_data)} entries")
        
        # Check if all values are 0 (formulas not calculated)
        if all(d.get('mrp') == 0 and d.get('offer') == 0 for d in self.summary_data):
            print("\n⚠️  WARNING: All MRP and Offer values are 0.")
            print("   This usually means formulas haven't been calculated.")
            print("   Please:")
            print("   1. Open this Excel file in Microsoft Excel")
            print("   2. Click 'Enable Content' or 'Update Links' when prompted")
            print("   3. Save the file (Ctrl+S / Cmd+S)")
            print("   4. Run create-summary again")
        
        return True
    
    def _collect_sheet_totals(self):
        """Collect total values from all data sheets"""
        for sheet_name in self.workbook.sheetnames:
            # Skip summary sheets
            if self._is_summary_sheet(sheet_name):
                continue

            sheet = self.workbook[sheet_name]
            totals = self._extract_totals(sheet, sheet_name)

            if totals:
                self.summary_data.append(totals)
                print(f"    {sheet_name}: MRP={totals.get('mrp', 'N/A')}, Offer={totals.get('offer', 'N/A')}")

    def _is_summary_sheet(self, sheet_name: str) -> bool:
        """Check if sheet is a summary sheet"""
        return any(summary.lower() in sheet_name.lower() for summary in self.config.SUMMARY_SHEET_NAMES)

    def _extract_totals(self, sheet, sheet_name: str) -> Optional[Dict]:
        """
        Extract total values from a sheet

        Args:
            sheet: openpyxl worksheet
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet name and totals
        """
        # Find TOTAL VALUE row
        total_row_idx = self._find_total_row(sheet)

        if not total_row_idx:
            return None
        
        # Try to get values from data_workbook first (for already calculated formulas)
        mrp_value = None
        offer_value = None
        
        if self.data_workbook and sheet_name in self.data_workbook.sheetnames:
            data_sheet = self.data_workbook[sheet_name]
            mrp_value = data_sheet.cell(row=total_row_idx, column=7).value
            offer_value = data_sheet.cell(row=total_row_idx, column=9).value
        
        # If values are still None, formulas haven't been calculated by Excel yet
        if mrp_value is None or offer_value is None:
            print(f"    ⚠️  {sheet_name}: Formulas not calculated. Please open Excel file, let formulas calculate, and Save.")
            # Return 0 values so summary is created but with placeholder values
            mrp_value = mrp_value or 0
            offer_value = offer_value or 0
        if mrp_value is None:
            mrp_value = 0
            # Sum column G (Total Amount = F * E) from row 4 to total_row - 1
            # Since column G has formulas, calculate from F (MRP) * E (Qty)
            for row_idx in range(4, total_row_idx):
                mrp = sheet.cell(row=row_idx, column=6).value  # Column F = MRP
                qty = sheet.cell(row=row_idx, column=5).value  # Column E = Qty
                if isinstance(mrp, (int, float)) and isinstance(qty, (int, float)):
                    mrp_value += mrp * qty
        
        if offer_value is None:
            offer_value = 0
            # Sum column I (Total Offer Price = H * E) from row 4 to total_row - 1
            # Since column I has formulas, calculate from H (Offer Price) * E (Qty)
            for row_idx in range(4, total_row_idx):
                offer = sheet.cell(row=row_idx, column=8).value  # Column H = Offer Price
                qty = sheet.cell(row=row_idx, column=5).value  # Column E = Qty
                if isinstance(offer, (int, float)) and isinstance(qty, (int, float)):
                    offer_value += offer * qty
        
        # Convert to numeric if needed
        if mrp_value and isinstance(mrp_value, str):
            try:
                mrp_value = float(mrp_value.replace('₹', '').replace('$', '').replace(',', '').strip())
            except:
                mrp_value = None
        elif isinstance(mrp_value, (int, float)):
            mrp_value = float(mrp_value)
        
        if offer_value and isinstance(offer_value, str):
            try:
                offer_value = float(offer_value.replace('₹', '').replace('$', '').replace(',', '').strip())
            except:
                offer_value = None
        elif isinstance(offer_value, (int, float)):
            offer_value = float(offer_value)

        return {
            'sheet_name': sheet_name,
            'mrp': mrp_value,
            'offer': offer_value,
            'row': total_row_idx
        }
    
    def _find_total_row(self, sheet) -> Optional[int]:
        """
        Find row containing TOTAL VALUE label

        Args:
            sheet: openpyxl worksheet

        Returns:
            Row number or None
        """
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for cell in row:
                if cell and isinstance(cell, str):
                    cell_upper = cell.upper().strip()
                    # Check for TOTAL VALUE keywords
                    if any(keyword.upper() in cell_upper for keyword in self.config.TOTAL_VALUE_KEYWORDS):
                        return row_idx
        return None
    
    def _extract_numeric_value(self, row_values: tuple, value_type: str) -> Optional[float]:
        """
        Extract numeric value from row

        Args:
            row_values: Tuple of row values
            value_type: 'mrp' or 'offer'

        Returns:
            Numeric value or None
        """
        # Look for numeric values in the row
        # Strategy: Take the last 2-3 numeric values (typically MRP and Offer)
        numeric_values = []

        for val in row_values:
            if val is None:
                continue

            # Try to convert to number
            try:
                if isinstance(val, (int, float)):
                    numeric_values.append(float(val))
                elif isinstance(val, str):
                    # remove currency symbols and commas
                    cleaned = val.replace('₹', '').replace('$', '').replace(',', '').strip()
                    if cleaned:
                        numeric_values.append(float(cleaned))
            except:
                continue

        # Return appropriate value
        if value_type == 'mrp' and len(numeric_values) >= 2:
            return numeric_values[-2]  # Second last is MRP
        elif value_type == 'offer' and len(numeric_values) >= 1:
            return numeric_values[-1]  # Last is Offer Price
        
        return None
    
    def _get_summary_sheet(self):
        """Get or create summary sheet"""
        # Check if summary sheet exists
        for sheet_name in self.workbook.sheetnames:
            if self._is_summary_sheet(sheet_name):
                # Clear existing content (except headers)
                sheet = self.workbook[sheet_name]
                return sheet
            
        # Create new summary sheet
        summary_sheet = self.workbook.create_sheet("SUMMARY", 0)
        return summary_sheet
    
    def _write_summary(self, sheet):
        """
        Write summary data to sheet
        
        Args:
            sheet: Summary worksheet
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Clear existing data first
        sheet.delete_rows(1, sheet.max_row)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        
        # Define styles
        header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers (row 1)
        headers = ['Sr. No', 'Sheet Name', 'MRP', 'OFFER PRICE']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Write data rows
        total_mrp = 0
        total_offer = 0
        for row_idx, data in enumerate(self.summary_data, start=2):
            # Sr. No
            cell = sheet.cell(row=row_idx, column=1)
            cell.value = row_idx - 1
            cell.alignment = center_align
            cell.border = border
            
            # Sheet Name
            cell = sheet.cell(row=row_idx, column=2)
            cell.value = data['sheet_name']
            cell.border = border
            
            # MRP
            mrp_val = data.get('mrp') or 0
            cell = sheet.cell(row=row_idx, column=3)
            cell.value = mrp_val
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            total_mrp += mrp_val if isinstance(mrp_val, (int, float)) else 0
            
            # Offer Price
            offer_val = data.get('offer') or 0
            cell = sheet.cell(row=row_idx, column=4)
            cell.value = offer_val
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            total_offer += offer_val if isinstance(offer_val, (int, float)) else 0
        
        # Add TOTAL MRP row
        total_row = len(self.summary_data) + 2
        cell = sheet.cell(row=total_row, column=2)
        cell.value = 'TOTAL MRP'
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        
        cell = sheet.cell(row=total_row, column=3)
        cell.value = total_mrp
        cell.number_format = '#,##0'
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        cell = sheet.cell(row=total_row, column=4)
        cell.value = total_offer
        cell.number_format = '#,##0'
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        # Add FINAL OFFER VALUE row
        final_row = total_row + 2
        cell = sheet.cell(row=final_row, column=2)
        cell.value = 'FINAL OFFER VALUE ( INCL GST )'
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        
        cell = sheet.cell(row=final_row, column=3)
        cell.value = total_offer
        cell.number_format = '#,##0'
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right')
        cell.border = border