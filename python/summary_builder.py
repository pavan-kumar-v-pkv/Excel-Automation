"""
Summary Builder Module
Generates summary sheets from data sheets
"""

from typing import Dict, List, Optional, Tuple
import openpyxl.styles

class SummaryBuilder:
    """
    Build summary sheet from data sheets
    """

    def __init__(self, workbook, config=None):
        """
        Initialize summary builder

        Args:
            workbook: openpyxl workbook object
            config: Configuration object
        """
        from .config import Config

        self.workbook = workbook
        self.config = config or Config()
        self.summary_data = []

    def build_summary(self) -> bool:
        """
        Build or update summary sheet

        Returns:
            True if successful
        """
        print(f"\n📊 Building Summary Sheet...")

        # Collect data from all sheets
        self._collect_sheet_totals()

        if not self.summary_data:
            print("    No data found to summarise")
            return False
        
        # get or create summary sheet
        summary_sheet = self._get_summary_sheet()

        # Write summary data
        self._write_summary(summary_sheet)

        print(f"Summary sheet update with {len(self.summary_data)} entries")
        return True
    
    def _collect_sheet_totals(self):
        """Collect total values from all data sheets"""
        for sheet_name in self.workbook.sheetnames:
            # Skip summary sheets
            if self._is_summary_sheet(sheet_name):
                continue

            sheet = self.workbook[sheet_name]
            totals = self._extract_totals(sheet, sheet_name)

            if totals:
                self.summary_data.append(totals)
                print(f"    {sheet_name}: MRP={totals.get('mrp', 'N/A')}, Offer={totals.get('offer', 'N/A')}")

    def _is_summary_sheet(self, sheet_name: str) -> bool:
        """Check if sheet is a summary sheet"""
        return any(summary.lower() in sheet_name.lower() for summary in self.config.SUMMARY_SHEET_NAMES)

    def _extract_totals(self, sheet, sheet_name: str) -> Optional[Dict]:
        """
        Extract total values from a sheet

        Args:
            sheet: openpyxl worksheet
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet name and totals
        """
        # Find TOTAL VALUE row
        total_row_idx = self._find_total_row(sheet)

        if not total_row_idx:
            return None
        
        # Extract values from total row
        row_values = list(sheet.iter_rows(
            min_row=total_row_idx,
            max_row=total_row_idx,
            values_only=True
        ))[0]

        # Find MRP and Offer Price Columns
        mrp_value = self._extract_numeric_value(row_values, 'mrp')
        offer_value = self._extract_numeric_value(row_values, 'offer')

        return {
            'sheet_name': sheet_name,
            'mrp': mrp_value,
            'offer': offer_value,
            'row': total_row_idx
        }
    
    def _find_total_row(self, sheet) -> Optional[int]:
        """
        Find row containing TOTAL VALUE label

        Args:
            sheet: openpyxl worksheet

        Returns:
            Row number or None
        """
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for cell in row:
                if cell and isinstance(cell, str):
                    cell_upper = cell.upper().strip()
                    # Check for TOTAL VALUE keywords
                    if any(keyword.upper() in cell_upper for keyword in self.config.TOTAL_VALUE_KEYWORDS):
                        return row_idx
        return None
    
    def _extract_numeric_value(self, row_values: tuple, value_type: str) -> Optional[float]:
        """
        Extract numeric value from row

        Args:
            row_values: Tuple of row values
            value_type: 'mrp' or 'offer'

        Returns:
            Numeric value or None
        """
        # Look for numeric values in the row
        # Strategy: Take the last 2-3 numeric values (typically MRP and Offer)
        numeric_values = []

        for val in row_values:
            if val is None:
                continue

            # Try to convert to number
            try:
                if isinstance(val, (int, float)):
                    numeric_values.append(float(val))
                elif isinstance(val, str):
                    # remove currency symbols and commas
                    cleaned = val.replace('₹', '').replace('$', '').replace(',', '').strip()
                    if cleaned:
                        numeric_values.append(float(cleaned))
            except:
                continue

        # Return appropriate value
        if value_type == 'mrp' and len(numeric_values) >= 2:
            return numeric_values[-2]  # Second last is MRP
        elif value_type == 'offer' and len(numeric_values) >= 1:
            return numeric_values[-1]  # Last is Offer Price
        
        return None
    
    def _get_summary_sheet(self):
        """Get or create summary sheet"""
        # Check if summary sheet exists
        for sheet_name in self.workbook.sheetnames:
            if self._is_summary_sheet(sheet_name):
                # Clear existing content (except headers)
                sheet = self.workbook[sheet_name]
                return sheet
            
        # Create new summary sheet
        summary_sheet = self.workbook.create_sheet("SUMMARY", 0)
        return summary_sheet
    
    def _write_summary(self, sheet):
        """
        Write summary data to sheet
        
        Args:
            sheet: Summary worksheet
        """
        # Write headers
        headers = ['Sheet Name', 'Total MRP', 'Total Offer Price']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Write data
        for row_idx, data in enumerate(self.summary_data, start=2):
            sheet.cell(row=row_idx, column=1).value = data['sheet_name']
            sheet.cell(row=row_idx, column=2).value = data.get('mrp')
            sheet.cell(row=row_idx, column=3).value = data.get('offer')
        
        # Auto-size columns
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width