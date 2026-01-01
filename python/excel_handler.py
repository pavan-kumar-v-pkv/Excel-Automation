"""
Excel Handler module
Handles reading from and writing to Excel files
"""

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple
from pathlib import Path

from .pdf_parser import normalize_sku

class ExcelHandler:
    """Handle Excel file operations"""

    def __init__(self, excel_path: str, config=None):
        """
        Initialize Excel handler
        Args:
            excel_path: Path to Excel file
            config: Configuration object
        """
        from .config import Config

        self.excel_path = excel_path
        self.config = config or Config()
        self.workbook = None
        self.sheet_sku_map = {} # {sheet_name: {row: sku}}

    def open(self):
        """Open Excel workbook"""
        print(f"\n Opening Excel: {self.excel_path}")
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path)
            print(f"    Found {len(self.workbook.sheetnames)} sheets")
        except Exception as e:
            print(f"    Error opening Excel: {str(e)}")
            raise

    def close(self):
        """Close workbook without saving"""
        if self.workbook:
            self.workbook.close()

    def save(self, output_path: str = None):
        """
        Save workbook
        Args:
            output_path: optional different path to save
        """
        save_path = output_path or self.excel_path
        print(f"\n Saving Excel: {save_path}")
        try:
            self.workbook.save(save_path)
            print("    Save successful")
        except Exception as e:
            print(f"    Error saving Excel: {str(e)}")
            raise

    def get_data_sheets(self) -> List[str]:
        """
        Get list of data sheets (excluding summary sheets)
        Returns:
            List of sheet names
        """
        if not self.workbook:
            return []
        
        data_sheets = []
        for sheet_name in self.workbook.sheetnames:
            # Skip summary sheets
            if any(summary.lower() in sheet_name.lower() for summary in self.config.SUMMARY_SHEET_NAMES):
                continue
            data_sheets.append(sheet_name)
        return data_sheets
    
    def scan_skus(self) -> Dict[str, List[Tuple[int, str]]]:
        """
        Scan all sheets for SKU codes

        Returns:
            Dict mapping sheet names to list of (row, sku) tuples
        """
        print(f"\n Scanning sheets for SKU codes...")
        sku_locations = {}
        data_sheets = self.get_data_sheets()

        for sheet_name in data_sheets:
            sheet = self.workbook[sheet_name]
            skus = self._scan_sheet_for_skus(sheet)
            if skus:
                sku_locations[sheet_name] = skus
                print(f"    {sheet_name}: Found {len(skus)} SKUs")

        self.sheet_sku_map = sku_locations

        all_excel_skus = set()
        for sheet, skus in sku_locations.items():
            for _, _, sku_norm in skus:
                all_excel_skus.add(sku_norm)
        print("SKUs in Excel (normalized):")
        for sku in sorted(all_excel_skus):
            print(sku)
        return sku_locations
    
    def _scan_sheet_for_skus(self, sheet) -> List[Tuple[int, str]]:
        """
        Scan a single sheet for SKU codes
        Args:
            sheet: openpyxl worksheet object
        Returns:
            List of (row_number, sku_code) tuples
        """
        skus = []
        code_col_idx = None

        # Find CODE column
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=False):
            for cell in row:
                if cell.value and self._is_code_header(str(cell.value)):
                    code_col_idx = cell.column
                    break
            if code_col_idx:
                break

        if not code_col_idx:
            return skus
        
        # Scan for SKUs in CODE column
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) >= code_col_idx:
                sku_value = row[code_col_idx - 1]
                if sku_value and self._is_valid_sku(str(sku_value)):
                    sku_raw = str(sku_value).strip()
                    sku_norm = normalize_sku(sku_raw)
                    skus.append((row_idx, sku_raw, sku_norm))

        return skus
    
    def _is_code_header(self, header: str) -> bool:
        """Check if cell is a CODE column header"""
        header_upper = header.upper().strip()
        return any(name.upper() in header_upper for name in self.config.CODE_COLUMN_NAMES)

    def _is_valid_sku(self, sku: str) -> bool:
        """Validate SKU code"""
        sku = sku.strip()
        if len(sku) < 3:
            return False
        if sku.upper() in ['TOTAL', 'SUBTOTAL', 'GRAND TOTAL']:
            return False
        return True

    def insert_images(self, image_paths: Dict[str, str]) -> int:
        """
        Insert images into Excel sheets

        Args:
            image_paths: Dictionary mapping SKU to image file path

        Returns:
            Number of images inserted
        """
        print(f"\n Inserting images into Excel...")
        inserted_count = 0

        for sheet_name, skus in self.sheet_sku_map.items():
            sheet = self.workbook[sheet_name]

            # Find image column
            image_col_idx = self._find_image_column(sheet)

            if not image_col_idx:
                print(f"    {sheet_name}: No IMAGE column found, skipping")
                continue

            # Insert images for each SKU
            for row_num, sku_raw, sku_norm in skus:
                if sku_norm in image_paths:
                    print(f"Will insert image for SKU: {sku_raw} (normalized: {sku_norm})")
                    success = self._insert_image_at_cell(
                        sheet,
                        row_num,
                        image_col_idx,
                        image_paths[sku_norm]
                    )
                    if success:
                        inserted_count += 1
                        print(f"      ✓ Inserted at row {row_num}, col {image_col_idx}")
                else:
                    print(f"NO IMAGE FOR SKU: {sku_raw} (normalized: {sku_norm})")

            print(f"    {sheet_name}: Inserted {inserted_count} images so far")
        print(f"Total images inserted: {inserted_count}")
        return inserted_count
    
    def _find_image_column(self, sheet) -> Optional[int]:
        """
        Find the Image column in the sheet

        Args:
            sheet: openpyxl worksheet object
        Returns:
            Column index of IMAGE column or None
        """

        # Search in first 20 rows for header
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=False):
            for cell in row:
                if cell.value and self._is_image_header(str(cell.value)):
                    return cell.column
                
        return None
    
    def _is_image_header(self, header: str) -> bool:
        """Check if cell is an IMAGE column header"""
        header_upper = header.upper().strip()
        return any(name.upper() in header_upper for name in self.config.IMAGE_COLUMN_NAMES)
    
    def _insert_image_at_cell(self, sheet, row: int, col: int, image_path: str) -> bool:
        """
        Insert image at specified cell

        Args:
            sheet: openpyxl worksheet object
            row: Row number
            col: Column number
            image_path: Path to image file

        Returns:
            True if inserted successfully, False otherwise
        """
        try:
            # Create Excel image object
            img = XLImage(image_path)
            
            # Resize to fit cell (assuming standard row height)
            img.width = self.config.IMAGE_TARGET_SIZE[0]
            img.height = self.config.IMAGE_TARGET_SIZE[1]
            
            # Get cell reference
            cell_ref = f"{get_column_letter(col)}{row}"
            
            # Anchor image to cell
            img.anchor = cell_ref
            
            # Add image to sheet
            sheet.add_image(img)
            
            # Adjust row height to fit image
            sheet.row_dimensions[row].height = self.config.IMAGE_TARGET_SIZE[1] * 0.75
            
            return True
            
        except Exception as e:
            print(f"      Error inserting image at {get_column_letter(col)}{row}: {str(e)}")
            return False
            

            
                